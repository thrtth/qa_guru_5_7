import csv
import time
import os.path
import zipfile

import requests
import xlrd
from selenium import webdriver
from selene import browser
from pypdf import PdfReader
from openpyxl import load_workbook


PROJECT_ROOT_PATH = os.path.dirname(os.path.abspath(__file__))
RESOURCES_PATH = os.path.join(PROJECT_ROOT_PATH, 'resources')
TMP_PATH = os.path.join(PROJECT_ROOT_PATH, 'tmp')


def test_csv():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    with open(os.path.join(RESOURCES_PATH, 'eggs.csv'), 'w') as csvfile:
        csvwriter = csv.writer(csvfile, delimiter=',')
        csvwriter.writerow(['Anna', 'Pavel', 'Peter'])
        csvwriter.writerow(['Alex', 'Serj', 'Yana'])

    with open(os.path.join(RESOURCES_PATH, 'eggs.csv')) as csvfile:
        csvreader = csv.reader(csvfile)
        csv_row_list = []
        for row in csvreader:
            print(row)
            csv_row_list.append(row)
    assert csv_row_list[2] == ['Alex', 'Serj', 'Yana']


def test_dowload_file():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь к tmp

    options = webdriver.ChromeOptions()
    prefs = {
        "download.default_directory": TMP_PATH,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)

    browser.config.driver_options = options

    browser.open("https://github.com/pytest-dev/pytest")
    browser.element(".d-none .Button-label").click()
    browser.element('[data-open-app="link"]').click()

    time.sleep(5)
    size = os.path.getsize(os.path.join(TMP_PATH, 'pytest-main.zip'))
    assert size == 1564388


def test_downloaded_file_size():
    # TODO сохранять и читать из tmp, использовать универсальный путь
    url = 'https://selenium.dev/images/selenium_logo_square_green.png'
    r = requests.get(url)
    with open(os.path.join(TMP_PATH, 'selenium_logo.png'), 'wb') as file:
        file.write(r.content)
    size = os.path.getsize(os.path.join(TMP_PATH, 'selenium_logo.png'))
    assert size == 30803


def test_pdf():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    reader = PdfReader(os.path.join(RESOURCES_PATH, 'docs-pytest-org-en-latest.pdf'))
    number_of_pages = len(reader.pages)
    page = reader.pages[0]
    text = page.extract_text()
    print(page)
    print(number_of_pages)
    print(text)
    assert number_of_pages == 412


def test_xlrd():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    book = xlrd.open_workbook(os.path.join(RESOURCES_PATH, 'file_example_XLS_10.xls'))
    print(f'Количество листов {book.nsheets}')
    print(f'Имена листов {book.sheet_names()}')
    sheet = book.sheet_by_index(0)
    print(f'Количество столбцов {sheet.ncols}')
    print(f'Количество строк {sheet.nrows}')
    print(f'Пересечение строки 9 и столбца 1 = {sheet.cell_value(rowx=0, colx=1)}')
    # печать всех строк по очереди
    for rx in range(sheet.nrows):
        print(sheet.row(rx))
    assert sheet.cell_value(rowx=9, colx=1) == 'Vincenza'


def test_xlsx():
    # TODO оформить в тест, добавить ассерты и использовать универсальный путь
    workbook = load_workbook(os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    assert sheet.cell(row=3, column=2).value == 'Mara'


def test_zip():
    file_dir = os.listdir(RESOURCES_PATH)
    with zipfile.ZipFile('resources.zip', mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        for file in file_dir:
            add_file = os.path.join(RESOURCES_PATH, file)
            zf.write(add_file)

    with zipfile.ZipFile('resources.zip', mode='a') as zf:
        number_of_files = len(zf.infolist())
        for file in zf.infolist():
            name = os.path.basename(file.filename)
            size = file.file_size
            print(name, size)

            assert number_of_files == 4
            assert size == os.path.getsize(os.path.join(RESOURCES_PATH, name))
